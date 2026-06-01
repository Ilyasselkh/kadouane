# -*- coding: utf-8 -*-

from odoo import models, fields, api
import openpyxl
import base64
from io import BytesIO
from odoo import exceptions 
import collections
import math


class Kadouane(models.Model):
    _name = 'kadouane.kadouane'
    _description = 'kadouane.kadouane'
    _inherit = [ 'mail.thread', 'mail.activity.mixin']
    _order = "create_date desc"


    referenceID = fields.Char( readonly=True, required=True, copy=False, default='Nouvelle demande') 
    name = fields.Char(string="Désignation",compute="get_name_calendar", tracking=True)

    name = fields.Char(string="Designation", compute="get_name_calendar", search="_search_name", tracking=True)

    @api.depends('client_id')
    def get_name_calendar(self):
        for rec in self:
                rec.name = rec.client_id.name

    def _search_name(self, operator, value):
        return [('client_id.name', operator, value)]

    
    type = fields.Selection([('COLIS', 'COLIS'),('PALETTE', 'PALETTE'),('TC', 'TC')] ,string="Type d'expidition",tracking=True)
    typedmd = fields.Selection([('Normal', 'Normal'),('Express', 'Express'),('exceptionnelle', 'exceptionnelle'),('projet', 'projet')] ,string="Type de la demande",tracking=True)
    filexlsx = fields.Binary(string="File", attachment=True)
    filexlsx_filename = fields.Char(default="",tracking=True)
    statusFichierglobal = fields.Char(default="NOK")
    statusFichier = fields.Char(default="A telecharger", string="Status des colonnes")
    statusFichierOrgCOm = fields.Char(default="A telecharger", string="l'organisation COM")
    statusFichierClient = fields.Char(default="A telecharger", string="Numero client")
    statusFichierdevis = fields.Char(default="A telecharger", string="Devis")
    totalvfactures  = fields.Float()
    devisFactures  = fields.Char()
    nbrFactures =  fields.Integer(compute='_get_count_rec_nbr_factures',)


    #typetransit = fields.Char(string="Type de transit", compute="get_client_data", store="True")
    typetransit = fields.Char(string="Type de transit",tracking=True)
    transitvalue = fields.Float(string="valeur du dossier",tracking=True)
    transitvalueExp = fields.Float(string="valeur des couts exceptionnels",tracking=True)
    transitvalueokpayment  = fields.Boolean(string="OK pour paiement",tracking=True)

    state = fields.Selection([('nouvelle', 'Nouvelle demande'),('transport', 'Attente Transport'),('picking', 'Attente picking'), ('facturation', 'Attente Facturation'), ('douane', 'Attente Transit'),  ('Cloture', 'Attente DUM'), ('done', 'Attente F imputation'), ('mainlevee', 'Attente main levee'), ('archive', 'Archive'), ('corbeille', 'Corbeille')], default ='nouvelle' ,string="Status",tracking=True)
    stateShip = fields.Selection([ ('exp', 'Attente d\'expedition'),('expdone', 'Expédié')] ,string="Status",tracking=True)
    closedmd =  fields.Char(compute="_get_dossier_closestatus")


    transitaire = fields.Selection([('Gazet', 'Gazet'),('TIMAR','TIMAR')], default ='TIMAR',tracking=True)
    transitairereq  = fields.Char(compute="_is_tra_req", default="F") 
    @api.depends('state','stateShip','client_id')
    def _is_tra_req(self):
        for rec in self:
            if rec.arnetwork == "True" or rec.state in ['douane', 'Cloture', 'done','mainlevee','archive']:
                rec.transitairereq = "T"
            else :
                rec.transitairereq = "F"

    allBomsAreOk  = fields.Boolean(default="False")
    customsvaluefile = fields.Binary(string="Customs value", attachment=True)
    customsvalue = fields.Float(string="Customs value",tracking=True)

    

    client_id = fields.Many2one('kadouane.client', string='client',tracking=True)
    incoterm = fields.Char(string="Incoterm", compute="get_client_data", store="True")
    shipto = fields.Char(string="Ship to" , compute="get_client_data", store="True")
    billto = fields.Char(string="Bill to" , compute="get_client_data", store="True")
    picking =  fields.Datetime(string="DH Picking planifier",tracking=True)
    arnetwork = fields.Char(string="AR network?" , compute="get_client_data", store="True")
    transitmanagement = fields.Boolean(string="Transit management",compute="get_client_data", store="True")


    ismanager = fields.Boolean(string="ismanager", compute='get_user_grp')

    @api.depends('ismanager')
    def get_user_grp(self):
        res_user = self.env['res.users'].search([('id', '=', self._uid)])
        if res_user.has_group('kadouane.group_resadv') or res_user.has_group('kadouane.group_log_manager'):
            self.ismanager = True
        else:
            self.ismanager = False

    dn = fields.Integer()
    archivedone  = fields.Boolean(string="Dossier archivé",tracking=True)

    archiveclosed  = fields.Boolean(string="Dossier archivé",tracking=True, compute="")

    @api.depends('archivedone', 'transportvalueokpayment','transitvalueokpayment')
    def get_checkifclosed(self):
        for rec in self:
                    if rec.archivedone == True and rec.transportvalueokpayment == True and rec.transitvalueokpayment == True:
                        rec.archiveclosed = True
                    else:
                        rec.archiveclosed = False


    typetransport = fields.Selection([('ROUTIER', 'ROUTIER'),('AERIEN', 'AERIEN'),('MARITIME', 'MARITIME')] ,string="Type du transport",tracking=True)

    valeurTransport = fields.Float(string="Valuer transport",tracking=True)
    CURTransport = fields.Selection([('MAD', 'MAD'),('EUR', 'EUR'),('USD', 'USD')],tracking=True )
    transportvalueokpayment  = fields.Boolean(string="OK pour paiement",tracking=True)

    transport_id = fields.Many2one('kadouane.transporteur', string='transporteur',tracking=True)


    transporteurReqNvlDmd = fields.Char(compute="_is_trans_req_nvl_dmd", default="F")
    transporteurReqNvlDmdred = fields.Char(compute="_is_trans_req_nvl_dmd", default="F")

    @api.depends('client_id', 'typedmd','state')
    def _is_trans_req_nvl_dmd(self):
        for rec in self:
            if (rec.typedmd == "Normal" and rec.incoterm == "FCA") or rec.state != "nouvelle":
                rec.transporteurReqNvlDmd = "T"
            else :
                rec.transporteurReqNvlDmd = "F"
            
            if rec.state == "nouvelle" or rec.state == "transport":
               rec.transporteurReqNvlDmdred = "F"
            else :
                rec.transporteurReqNvlDmdred = "R"

    


    nbrpalletsadv = fields.Integer(string="Nbr palettesa expidier",tracking=True)
    nbrpalletcolis = fields.Integer(string="Nbr colis a expidier",tracking=True)
    nbrpalletcolisadvok = fields.Char(default="F", compute="_get_nbrpalstatus")
    @api.depends("nbrpalletsadv","nbrpalletcolis" )
    def _get_nbrpalstatus(self):
        for rec in self:
            if rec.nbrpalletsadv < 1 and rec.nbrpalletcolis <1:
                rec.nbrpalletcolisadvok = "F"
            else :
                rec.nbrpalletcolisadvok = "T"


   # @api.constrains('nbrpalletsadv', 'nbrpalletcolis')
   # def _check_value(self):
   #     if self.nbrpalletsadv < 1 and nbrpalletcolis < 1 :
   #         raise exceptions.UserError('Merci de rensigner le nbr des palletes / colis')
        
    
    commentaireTransport = fields.Char(string="Commantaire",tracking=True)
    commentaireBL = fields.Char(string="Commantaire",tracking=True)
    commentairegen = fields.Char(string="Commantaire",tracking=True)
    commentaireinv = fields.Char(string="Commantaire",tracking=True)
    commentairetransaire = fields.Char(string="Commantaire",tracking=True)
    commentairetranautres = fields.Char(string="Commantaire",tracking=True)
    commentaireprepa = fields.Char(string="Commantaire PREPA",tracking=True)
    commentaireexp = fields.Char(string="Commantaire global EXP",tracking=True)
    commentaireAdv = fields.Char(string="Commantaire ADV",tracking=True)

    preparationdone  = fields.Boolean(string="Préparé",tracking=True)
    
    
    attachment_DN_ids = fields.Many2many('ir.attachment', relation="class_ir_attachments_rel", 
                                column1="class_id",
                                column2="attachment_id", string="les Bon de livraison",tracking=True)
    
    attachment_inv_ids = fields.Many2many('ir.attachment', relation="m2m_ir_invoices_rel", 
                                column1="m2m_id1",column2="attachment_id", string="Les factures",tracking=True)

    
    attachment_transport_ids = fields.Many2many('ir.attachment', relation="m2m_ir_transport_rel", 
                                column1="m2m_id2",
                                column2="attachment_id", string="Documents du transport",tracking=True)
    attachment_transitaire_ids = fields.Many2many('ir.attachment', relation="m2m_ir_transit_rel", 
                                column1="m2m_id3",
                                column2="attachment_id", string="Documents du transitaire",tracking=True)
    
    attachment_autres_ids = fields.Many2many('ir.attachment', relation="m2m_ir_autres_rel", 
                                column1="m2m_id4",
                                column2="attachment_id", string="Autres documents",tracking=True)
    
    
    
    dn_ids = fields.Many2one(comodel_name='kadouane.dn', string='dns',tracking=True)
    dn_ids_withattachement = fields.Many2one(comodel_name='kadouane.dnwithatachement', string='dns',tracking=True)

    
    attachment_docs_ids = fields.One2many('kadouane.autresdocs', 'docu_id', string='autres documents',tracking=True)
    dn_ids = fields.One2many('kadouane.dn', 'dn_id', string='DNs',tracking=True)
    dn_ids_withattachement = fields.One2many('kadouane.dnwithatachement', 'ids_withattachement', string='DNs',tracking=True)
    inv_ids = fields.One2many('kadouane.inv', 'inv_id', string='invoices',tracking=True)
    inv_refs = fields.One2many('kadouane.refexp', 'lines_kadouane_ids', string='Eclatement BOM',tracking=True)
    dechargengp = fields.One2many('kadouane.dechargeexportngp', 'lines_kadouane_ids', string='Decharge',tracking=True)
    remorque = fields.One2many('kadouane.remorque', 'lines_kadouane_ids', string='Expeditions',tracking=True)

    def _get_count_rec_nbr_factures(self):
        for rec in self:
            rec.nbrFactures = self.env['kadouane.inv'].search_count([('inv_id','=',self.id)])
          


    @api.constrains('filexlsx')
    def get_data(self):
        self.statusFichier = "A telecharger"
        self.statusFichierOrgCOm = "A telecharger"
        self.statusFichierClient = "A telecharger"
        self.statusFichierdevis = "A telecharger"
        if isinstance(self.filexlsx_filename, str):  
            if not self.filexlsx_filename.lower().endswith('.xlsx'):     # check if file pdf
                raise exceptions.ValidationError("le fichier n'est pas un XLSX")
            else:
                pass

    @api.depends("stateShip","state" )
    def _get_dossier_closestatus(self):
        for rec in self:
            if rec.stateShip == 'expdone' and rec.state == 'done':
                rec.closedmd = "T"
            else :
                rec.closedmd = "F" 


        
    def action_transport(self):
            if self.nbrpalletcolisadvok == "F":
                raise exceptions.UserError('Merci de rensigner le nbr des palletes / colis')
            if len(self.dn_ids_withattachement) < 1:
                raise exceptions.UserError('Merci de rensigner minumum un BL dans la zone "Bon de livraison"')
            if self.incoterm == "FCA" and self.typedmd == "Normal":
                self.state = 'picking'
            else:
                self.state = 'transport'

    def action_nouvelle(self):
        self.state = 'nouvelle'
        
    def action_supprimer(self):
        self.state = 'corbeille'

    def action_facturation(self):
        for rec in self:
            if rec.preparationdone == True :
                rec.state = 'facturation'
            else :
                raise exceptions.UserError('Merci de cocher Préparé dans la zone "INFORMATIONS EXPIDITION"')
           
    
    def action_picking(self):
        self.state = 'picking'

    def action_douane(self):
        if self.statusFichier != "fichier OK" or self.statusFichierOrgCOm != "fichier OK" or self.statusFichierClient != "fichier OK" or self.statusFichierdevis != "fichier OK":
            raise exceptions.UserError('cliquer sur importer dans la section "IMPORTATION DES DONNÉES SAP"')


        if len(self.inv_ids) < 1:
            raise exceptions.UserError('Merci de rensigner minumum une facture dans la zone "Factures"')  
        else:
            for invs in self.inv_ids:
                if invs.attachment_INV_ids1 == False :
                    tt = "la facture : " + invs.numeroFacture + " ne contient aucun fichier attaché"
                    raise exceptions.UserError(tt) 
              
            if self.transitmanagement == False :
                self.state = 'archive'
            else :
                self.state = 'douane'
            self.stateShip = 'exp'

    def action_exp(self):
        if self.preparationdone == False:
            raise exceptions.UserError('Merci valider la preparation dans la section "INFORMATION EXPIDTION"')
        self.stateShip = 'exp'

    def action_dedouane(self):
        if len(self.dechargengp) < 1:
            raise exceptions.UserError('Merci de calculer les NGP "Button calcule NGP dans la section INFORMATIONS TRANSITAIRE"')
        else:
            self.state = 'Cloture'

    def action_Cloture(self):
        if len(self.inv_ids) < 1:
            raise exceptions.UserError('Merci de rensigner minumum une facture dans la zone "Factures"')  
        else:
            for invs in self.inv_ids:
                if invs.dum == False or invs.dum == "" :
                    tt = "Merci de rensigner le numero de dum pour la facture : " + invs.numeroFacture
                    raise exceptions.UserError(tt)   
            self.state = 'done'


    def action_done(self):
       self.state = 'done'

    def action_expdone(self):
           # for rec in self:
            #    if rec.expidtionDate == False:
            #        raise exceptions.UserError('Merci de rensigner la date d\'expidtion')
            #    if rec.cmr == False or rec.cmr == "":
            #        raise exceptions.UserError('Merci de rensigner le numero de CMR')
            #    if rec.plaqueTransport == False or rec.plaqueTransport == "":
            #        raise exceptions.UserError('Merci de rensigner le numero de la plaque')
            #    if rec.telchauffeur == False or rec.telchauffeur == "":
            #        raise exceptions.UserError('Merci de rensigner le numero de tel du chauffeur')

            #    if rec.nbrpallets < 1 and rec.nbrcolis <1:
            #        raise exceptions.UserError('Merci de rensigner le nbr des palletes / colis expédié')
            #    else :

        if len(self.remorque) < 1:
            raise exceptions.UserError('Merci de rensigner minumum une expedition dans la zone "INFORMATIONS EXPIDITION"')  
        else:
            for invs in self.remorque:
                if invs.nbrpallets < 1 and  invs.nbrcolis < 1 :
                    tt = "Merci de rensigner le nbr colis/pallets expidé pour l\'expedition' : " + invs.plaqueTransport
                    raise exceptions.UserError(tt)   

        for rec in self:
            if rec.typetransport == 'AERIEN' or rec.typetransport ==  'MARITIME' :
                for invs in self.remorque:
                    if invs.poidsbrut > 0 and  invs.poidsnet > 0 and invs.volume > 0 :
                        self.stateShip = 'expdone'  
                    else:
                        tt = "Merci de rensigner le poids net&brut et le volume pour l\'expedition' : " + invs.plaqueTransport
                        raise exceptions.UserError(tt)
                        
            else:
                self.stateShip = 'expdone'


    def action_Cloture_done(self):

        if len(self.dechargengp) < 1:
            raise exceptions.UserError('erreur pas de facture')  
        else:
            for invs in self.dechargengp:
                if invs.numdecharge == False or invs.numdecharge == "" :
                    tt = "Merci de rensigner le/les numero de decharge pour la NGP' : " + invs.ngp_rm
                    raise exceptions.UserError(tt)   

        for rec in self:
            if rec.transportvalueokpayment == False:
                 raise exceptions.UserError('Merci de valider le cout du transport (Ok pour paiement) dans la section "INFORMATIONS TRANSPORT"')
            
            if rec.transitvalueokpayment == False:
                 raise exceptions.UserError('Merci de valider le cout et le cout expetionnel du transit  (Ok pour paiement) dans la section "INFORMATIONS TRANSITAIRE"')

            if rec.archivedone == True :
                if rec.transitvalue > 0 :
                    self.state = 'mainlevee'
                else :
                    raise exceptions.UserError('Merci de rensigner la valuer du dossier et la valeur des couts exceptionnels dans la zone "INFORMATIONS TRANSITAIRE"')
            else :
                raise exceptions.UserError('Merci de cocher dossier archivé dans la zone "INFORMATIONS TRANSITAIRE"')

            


        

    def save(self, cr, uid, ids, context):
        return True
    
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('referenceID', 'Nouvelle demande') in [False, 'New', 'Nouvelle demande']:
                vals['referenceID'] = self.env['ir.sequence'].next_by_code(
                    'kadouane.kadouane.sequence'
                ) or 'New'

        records = super().create(vals_list)
        return records

    @api.depends('client_id')
    def get_client_data(self):
        self.incoterm = self.client_id.incoterm
        self.shipto = self.client_id.shipto
        self.billto = self.client_id.billto
        self.arnetwork = self.client_id.arnetwork
        self.transitmanagement = self.client_id.transitmanagement




    def import_customer(self):
        self.statusFichier = "A telecharger"
        self.statusFichierOrgCOm = "A telecharger"
        self.statusFichierClient = "A telecharger"
        self.statusFichierdevis = "A telecharger"
        self.checkXLSfileFromSAP()
        
    def calculeNGP(self):

        #explosionBOM
        try:
            searchifrefsexistanddelete = self.env['kadouane.refexp'].search([('lines_kadouane_ids', '=', self.id)])
            for recds in searchifrefsexistanddelete:
                recds.unlink()

            searchifNGPsexistanddelete = self.env['kadouane.dechargeexportngp'].search([('lines_kadouane_ids', '=', self.id)])
            for recds in searchifNGPsexistanddelete:
                recds.unlink()

            wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.filexlsx)), read_only=True)
            ws = wb.active

            for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=22, values_only=True):

                search = self.env['kadouane.bom'].search([('article_fg', '=', record[11])])
                
                if search:
                    for a in search:
                        
                        self.env['kadouane.refexp'].create({                           
                                'ngp_FG': a.ngp_FG,
                                'um_FG': a.um_FG,
                                'article_rm': a.article_rm,
                                'designation_rm': a.designation_rm,
                                'ngp_rm': a.ngp_rm,
                                'um_rm': a.um_rm,
                                'BOMqty': a.BOMqty,
                                'devis': record[0],
                                'clientshipto': record[1],
                                'numerofacture': record[4],
                                'numeroBl': record[9],
                                'article_fg': record[11],
                                'designation_fg': record[12],
                                'qtyfacture': record[14],
                                'valfacture': record[15],
                                'lines_kadouane_ids' : self.id
                        })

                else:
                    self.env['kadouane.refexp'].create({                           
                            'ngp_FG': '',
                            'um_FG': '',
                            'article_rm': '',
                            'designation_rm': '',
                            'ngp_rm': '',
                            'um_rm': '',
                            'BOMqty': '',
                            'devis': record[0],
                            'clientshipto': record[1],
                            'numerofacture': record[4],
                            'numeroBl': record[9],
                            'article_fg': record[11],
                            'designation_fg': record[12],
                            'qtyfacture': record[14],
                            'valfacture': record[15],
                            'lines_kadouane_ids' : self.id
                    })
                
        except:
            raise exceptions.ValidationError("fichier importé n'est pas valide")

        #fin explosion BOM

        searchifNGPsexistanddelete = self.env['kadouane.dechargeexportngp'].search([('lines_kadouane_ids', '=', self.id)])
        for recds in searchifNGPsexistanddelete:
            recds.unlink()

        search1 = self.env['kadouane.inv'].search([('inv_id', '=', self.id), ('regime', '=', False) ])
        if search1:
            err = []
            for a in search1:
                err.append(a.numeroFacture)
            raise exceptions.ValidationError("Il faut selectionner le regime pour les factures :"+str(err))
        
        search1 = self.env['kadouane.refexp'].search([('lines_kadouane_ids', '=', self.id), ('article_rm', '=', '')])
        if search1:
            err = []
            for a in search1:
                err.append(a.article_fg)
            raise exceptions.ValidationError("des references sans BOM :"+str(err))
            
        search1 = self.env['kadouane.refexp'].search([('lines_kadouane_ids', '=', self.id), ('BOMqty', '=', '')])
        if search1:
            err = []
            for a in search1:
                err.append(a.article_fg)
            raise exceptions.ValidationError("des references avec Qty BOM erroné :"+str(err))
        else:
            search1 = self.env['kadouane.refexp'].search([('lines_kadouane_ids', '=', self.id), ('BOMqty', '=', '0')])
            if search1:
                err = []
                for a in search1:
                    err.append(a.article_fg)
                raise exceptions.ValidationError("des references avec Qty BOM 0 :"+str(err))


        search1 = self.env['kadouane.refexp'].search([('lines_kadouane_ids', '=', self.id), ('ngp_rm', '=', '')])
        if search1:
            err = []
            for a in search1:
                err.append(a.article_fg)
            raise exceptions.ValidationError("des references sans NGP MP :"+str(err))
        
        search = self.env['kadouane.refexp'].search([('lines_kadouane_ids', '=', self.id)])
        ngps = []

        if search:
            for r in search:
                tmplist={}
                search2 = self.env['kadouane.inv'].search([('inv_id', '=', self.id), ('numeroFacture', '=', r.numerofacture) ])
                tmplist[str(search2.regime)+"##"+str(r.ngp_rm)+"##"+r.um_rm] = r.totalqty
                ngps.append(tmplist) 


                # sum the values with same keys
        
            counter = collections.Counter()
            for d in ngps:
                counter.update(d)
                
            ngps = dict(counter)

            for np in ngps:
                
                strsplit = str(np).split('##')


                self.env['kadouane.dechargeexportngp'].create({                           
                                'ngp_rm': strsplit[1],
                                'um_rm': strsplit[2],
                                'regim': strsplit[0],
                                'qty': ngps[np],
                                'lines_kadouane_ids' : self.id
                        })
        else:
            raise exceptions.ValidationError("erreur les ref ngp ne sont pas cree")

               

           

    def checkXLSfileFromSAP(self):
        
        if isinstance(self.filexlsx_filename, str):

            if self.filexlsx_filename.lower().endswith('.xlsx'):
                try:
                    self.statusFichier = "A telecharger"
                    self.statusFichierOrgCOm = "A telecharger"
                    self.statusFichierClient = "A telecharger"
                    self.statusFichierdevis = "A telecharger"

                    
                    searchiffactureexistanddelete = self.env['kadouane.inv'].search([('inv_id', '=', self.id)])
                    for recds in searchiffactureexistanddelete:
                        recds.unlink()
                    
                    searchifdnexistanddelete = self.env['kadouane.dn'].search([('dn_id', '=', self.id)])
                    for recds in searchifdnexistanddelete:
                        recds.unlink()


                    wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.filexlsx)), read_only=True)
                    ws = wb.active
                except:
                    raise exceptions.ValidationError("fichier importé n'est pas valide")
                
                #ws = wb['ImportationExport']
                for record in ws.iter_rows(min_row=0, max_row=1, min_col=None,max_col=22, values_only=True):
                    
                    if record[0] == "Devise document":
                        if record[0] != "Devise document" or record[1] != "Donneur d'ordre" or record[2] != "Payeur" or record[3] != "Nom 1" or record[4] != "Facture" or record[5] != "Créé par" or record[6] != "Créé le" or record[7] != "Heure" or record[8] != "Date de la facture" or record[9] != "Modèle de document" or record[10] != "Organis. commerciale" or record[11] != "Article" or record[12] != "Désignation" or record[13] != "Numéro d'article client" or record[14] != "Quantité facturée" or record[15] != "Valeur nette" or record[16] != "Valeur nette" or record[17] != "Devise statistique" or record[18] != "Poids net" or record[19] != "Poids brut" or record[20] != "Unité de poids" or record[21] != "WE-Client":
                            #raise exceptions.ValidationError(record[0])
                            raise exceptions.ValidationError("les colonnes du fichier importé n'est pas valide")
                        else:
                            self.statusFichier = "fichier OK"
                    elif record[0] == "Document Currency":
                        if record[0] != "Document Currency" or record[1] != "Sold-To Party" or record[2] != "Payer" or record[3] != "Name 1" or record[4] != "Billing Document" or record[5] != "Created By" or record[6] != "Created On" or record[7] != "Time" or record[8] != "Billing Date" or record[9] != "Reference Document" or record[10] != "Sales Organization" or record[11] != "Material" or record[12] != "Description" or record[13] != "Customer Material Number" or record[14] != "Billed Quantity" or record[15] != "Net value" or record[16] != "Net Value" or record[17] != "Statistics currency" or record[18] != "Net weight" or record[19] != "Gross weight" or record[20] != "Weight Unit" or record[21] != "WE-Customer":
                            raise exceptions.ValidationError("les colonnes du fichier importé n'est pas valide")
                        else:
                            self.statusFichier = "fichier OK"                            
                    else:
                        raise exceptions.ValidationError("les colonnes du fichier importé n'est pas valide")

                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=22, values_only=True):
                    if record[10] == "1051" :
                        self.statusFichierOrgCOm = "fichier OK"
                    else:
                        self.statusFichierOrgCOm = "A telecharger"
                        raise exceptions.ValidationError("L'organisation com dans le fichier n'est pas correcte")
                
                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=22, values_only=True):

                    if record[21] == self.shipto :
                        self.statusFichierClient = "fichier OK"
                    else:
                        self.statusFichierClient = "A telecharger"
                        raise exceptions.ValidationError("Le numero du client dans le fichier n'est pas correcte")



                firstdevis = ""
                firstdeviscaptured = 0
                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=22, values_only=True):
                    if firstdeviscaptured == 0:
                        firstdevis = record[0]
                        firstdeviscaptured = 1
                    if record[0] == firstdevis :
                        self.statusFichierdevis = "fichier OK"
                    else:
                        self.statusFichierdevis = "A telecharger"
                        raise exceptions.ValidationError("Les factures n'ont pas le meme devis")
                    self.devisFactures = firstdevis
                    
                
                bls1 = []
                factures = []
                #refss = []
                self.totalvfactures = 0
                
                for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=22, values_only=True):
                    tmplist={}
                    #tmplist1={}
                    tmplist2={}

                    tmplist[str(record[4])] = record[15]
                    factures.append(tmplist) 

                    tmplist2[str(record[9])] = str(record[4])
                    bls1.append(tmplist2)

                    #tmplist1[str(record[11])] = str(record[14])
                    #refss.append(tmplist1) 

                    self.totalvfactures =self.totalvfactures+ record[15]
               
                # sum the values with same keys
                
                counter = collections.Counter()
                for d in factures:
                    counter.update(d)
                    
                factures = dict(counter)

                # remove duplicate BL&INV
                seen = set()
                new_l = []
                for d in bls1:
                    t = tuple(d.items())
                    if t not in seen:
                        seen.add(t)
                        new_l.append(d)
                bls1 = new_l
                 # remove duplicate BL&INV

                
               
                
                #counter = collections.Counter()
                #for d in refss:
                #    counter.update(d)
                    
                #refss = dict(counter)

                #bls = list(dict.fromkeys(bls))

                
                
                for fct in factures:
                    
                    self.env['kadouane.inv'].create({            
                            'numeroFacture': fct,
                            'inv_id': self.id,
                            'valeurFacture': factures[fct],
                            'devis': firstdevis,
  
                    })
                for bl in bls1:
                    res = list(bl.keys())[0]
                    #raise exceptions.ValidationError(str(res))
                    self.env['kadouane.dn'].create({            
                            'numeroFacture': bl[res],
                            'dn_id': self.id,
                            'numeroBL': res,
  
                    })
                

            else:
                raise exceptions.ValidationError("Extension du fichier importé n'est pas valide")
        else:
            raise exceptions.ValidationError("Pas de fichier selectionné")
    
            

                
                
            # search if the customer exist else create
                
                #search = self.env['res.partner'].search([('name', '=', record[1]), ('customer_rank', '=', True)])
               
               
            #if not search:
                    
                #self.env['res.partner'].create({            
                #'ref': record[0],
                #'name': record[1],
                #'street': record[2],
                #'state_id': self.env['res.country.state'].search([
                #('name', '=', record[3])]).id,
                #'country_id': self.env['res.country'].search([
                #('code', '=', record[4])]).id,
                #'zip': record[5],
                #'phone': record[6],
                #'email': record[7],
                #'customer_rank': True
                #})


   
#     value = fields.Integer()
#     value2 = fields.Float(compute="_value_pc", store=True)
#     description = fields.Text()
#
#     @api.depends('value')
#     def _value_pc(self):
#         for record in self:
#             record.value2 = float(record.value) / 100



class kadouane(models.Model):
    _name = 'kadouane.client'
    _description = 'kadouane.client'




    name = fields.Char(tracking=True)
    shipto = fields.Integer(tracking=True)
    billto = fields.Integer(tracking=True)
    incoterm = fields.Char(tracking=True)
    typetransit = fields.Selection([('CESSION', 'CESSION'),('EXPORT', 'EXPORT'), ('TTC', 'TTC')] ,string="Type de transit",tracking=True)
    transportmanagement = fields.Boolean(tracking=True)
    transitmanagement = fields.Boolean(tracking=True)
    arnetwork = fields.Boolean(tracking=True)
    domicilation = fields.Selection([('SG', 'SAHAM BANQUE'),('Attijari', 'attijariwafa bank')] ,string="domiciliation",tracking=True)


class kadouane(models.Model):
    _name = 'kadouane.transporteur'
    _description = 'kadouane.transporteur'




    name = fields.Char(tracking=True)
    


#     value = fields.Integer()
#     value2 = fields.Float(compute="_value_pc", store=True)
#     description = fields.Text()
#
#     @api.depends('value')
#     def _value_pc(self):
#         for record in self:
#             record.value2 = float(record.value) / 100




class kadouane(models.Model):
    _name = 'kadouane.dn'
    _description = 'kadouane.dn'




    numeroBL = fields.Char(tracking=True)
    fildn_filename = fields.Char(tracking=True)
    numeroFacture = fields.Char(tracking=True)
    attachment_dn_ids1  = fields.Binary(string="Fichier PDF",attachment=True)
    #numerofacutre = fields.Char()
    dn_id = fields.Many2one('kadouane.kadouane', string='dn')


class kadouane(models.Model):
    _name = 'kadouane.dnwithatachement'
    _description = 'kadouane.dn'




    numeroBL = fields.Char(required=True,tracking=True)
    fildn_filename = fields.Char(required=True,tracking=True)
    attachment_dn_ids1  = fields.Binary(string="Fichier PDF",attachment=True,required=True)
    #numerofacutre = fields.Char()
    ids_withattachement = fields.Many2one('kadouane.kadouane', string='dn')


class kadouane(models.Model):
    _name = 'kadouane.refexp'
    _description = 'kadouane.refexp'



    article_fg = fields.Char(string="Article FG",tracking=True)
    designation_fg = fields.Char(tracking=True)
    ngp_FG = fields.Char(tracking=True)
    um_FG  = fields.Char(tracking=True)

    article_rm = fields.Char(tracking=True)
    designation_rm = fields.Char(tracking=True)
    ngp_rm = fields.Char(tracking=True)
    um_rm  = fields.Char(tracking=True)
    
    BOMqty  = fields.Float(tracking=True)

    numerofacture = fields.Char(tracking=True)
    numeroBl = fields.Char(tracking=True)
    qtyfacture = fields.Float(tracking=True)
    valfacture = fields.Float(tracking=True)
    devis  = fields.Char(tracking=True)
    clientshipto  = fields.Char(tracking=True)

    regim  = fields.Char(tracking=True)

    totalqty = fields.Float(compute="_calctotalqty",readonly=1,tracking=True)

    lines_kadouane_ids = fields.Many2one('kadouane.kadouane', string='reference')

    @api.depends('totalqty', 'qtyfacture', 'BOMqty' )
    def _calctotalqty(self):
        for rec in self:
            rec.totalqty = rec.BOMqty * rec.qtyfacture
            #raise exceptions.UserError(rec.valfacture)



class kadouane(models.Model):
    _name = 'kadouane.dechargeexportngp'
    _description = 'kadouane.dechargeexportngp'


    ngp_rm = fields.Char(tracking=True)
    um_rm  = fields.Char(tracking=True)
    regim  = fields.Char(tracking=True)
    numdecharge = fields.Char(string="Num compte decharge",tracking=True)
    qty  = fields.Float(tracking=True)
    
    lines_kadouane_ids = fields.Many2one('kadouane.kadouane', string='decharge')





    #lines_kadouane_ids = fields.Many2one('kadouane.kadouane', string='reference')

 


class kadouane(models.Model):
    _name = 'kadouane.remorque'
    _description = 'kadouane.remorque'


    poidsbrut = fields.Float(string="Poids Brut",tracking=True)
    poidsnet = fields.Float(string="Poids net",tracking=True)
    volume = fields.Float(tracking=True)
    plaqueTransport = fields.Char(string="Plaque",tracking=True)

    cmr = fields.Char(string="Numero de CMR",tracking=True)
    nbrpallets = fields.Integer(string="Nbr palettes expidié",tracking=True)
    nbrcolis = fields.Integer(string="Nbr colis expidié",tracking=True)
    nbrpalletcolisexpok = fields.Char(default="F", compute="_get_nbrpalexpstatus")
    telchauffeur = fields.Char(string="Tel chauffeur",tracking=True)
    numfactures = fields.Char(string="num des factures",tracking=True)
    expidtionDate =  fields.Datetime(string="DH expedition",tracking=True)

    lines_kadouane_ids = fields.Many2one('kadouane.kadouane', string='expeditions')

    @api.depends("nbrpallets","nbrcolis" )
    def _get_nbrpalexpstatus(self):
        for rec in self:
            if rec.nbrcolis < 1 and rec.nbrpallets <1:
                rec.nbrpalletcolisexpok = "F"
            else :
                rec.nbrpalletcolisexpok = "T"




    #lines_kadouane_ids = fields.Many2one('kadouane.kadouane', string='reference')



class kadouane(models.Model):
    _name = 'kadouane.bom'
    _description = 'kadouane.bom'
    _inherit = [ 'mail.thread', 'mail.activity.mixin']

    article_fg = fields.Char(string="Article FG",tracking=True)
    designation_fg = fields.Char(tracking=True)
    ngp_FG = fields.Char(tracking=True)
    um_FG  = fields.Char(tracking=True)

    article_rm = fields.Char(tracking=True)
    designation_rm = fields.Char(tracking=True)
    ngp_rm = fields.Char(tracking=True)
    um_rm  = fields.Char(tracking=True)
    
    BOMqty  = fields.Float(tracking=True)

    def action_nouvelle(self):
        self.BOMqty = 'nouvelle'
    
    
class kadouane(models.Model):
    _name = 'kadouane.inv'
    _description = 'kadouane.inv'




    numeroFacture = fields.Char(string="Num facture",tracking=True)
    #numerofacutre = fields.Char()
    inv_id = fields.Many2one('kadouane.kadouane', string='inv',tracking=True)
    regime = fields.Selection([('37', '37'),('22', '22')],tracking=True)
    dum = fields.Char(string="DUM",tracking=True)
    valeurFacture = fields.Float(tracking=True)
    attachment_INV_ids1  = fields.Binary(string="Fichier PDF",attachment=True)
    filinv_filename = fields.Char(string="Numero du compte utilisé",tracking=True)
    
    devis  = fields.Char(tracking=True)
    

class kadouaneautreDocs(models.Model):
    _name = 'kadouane.autresdocs'
    _description = 'kadouane.autresdocs'



  
    docu_id = fields.Many2one('kadouane.kadouane', string='dn')

    namedoc = fields.Char(string="Nom document",tracking=True)
    typedoc = fields.Selection([('transport', 'Transport'),('transitaire', 'transitaire'),('DUM', 'DUM'), ('Facture', 'Facture'), ('CMR', 'CMR'),  ('ficheIMp', 'fiche d\'imputation'), ('mainlevee', 'mainlevee'), ('photo', 'photo'),('autre', "autre")], required=True)
    attachment_fichier_ids  = fields.Binary(string="Fichier",attachment=True)






class kadouane(models.Model):
    _name = 'kadouane.categoryex'
    _description = 'kadouane.categoryex'



    title = fields.Char(tracking=True)
    color = fields.Char(tracking=True)
    fieldname = fields.Char(tracking=True)
    countnvl = fields.Integer(compute='_get_count_rec',default=0)
    counttransport = fields.Integer(compute='_get_count_rec',default=0)
    countdouane = fields.Integer(compute='_get_count_rec',default=0)
    countexp = fields.Integer(compute='_get_count_rec',default=0)
    countCloture = fields.Integer(compute='_get_count_rec',default=0)
    countpicking = fields.Integer(compute='_get_count_rec',default=0)
    countdone= fields.Integer(compute='_get_count_rec',default=0)
    countmainlevee = fields.Integer(compute='_get_count_rec',default=0)
    countFacturation = fields.Integer(compute='_get_count_rec',default=0)
    countArchive = fields.Integer(compute='_get_count_rec',default=0)
    countArchiveOK = fields.Integer(compute='_get_count_rec',default=0)

    statename = fields.Selection([('nouvelle', 'Nouvelle demande'),('transport', 'Attente Transport'),('picking', 'Attente picking'), ('facturation', 'Attente Facturation'), ('douane', 'Attente Transit'),  ('Cloture', 'Attente DUM'), ('done', 'Cloturé'), ('mainlevee', 'Attente main levee'),('exp', "Attente d'expedition"),('expdone', 'Expédié'), ('archive', 'Archive')])
   
    def get_stock_picking_action_picking_type(self):
        return self._get_action('kadouane.action_window1')
    
    def get_create_nouvelledemande_form_action(self):
        return {
            'res_model' : 'kadouane.kadouane',
            'type' :  'ir.actions.act_window',
            'view_mode' : 'form',
            'view_type' : 'form',
            'view_id' : self.env.ref("kadouane.view_dmd_form").id,
            'target' : 'self'
        }

    def _get_count_rec(self):
        for rec in self:
            rec.countnvl = self.env['kadouane.kadouane'].search_count([('state','=','nouvelle')])
            rec.countpicking = self.env['kadouane.kadouane'].search_count([('state','=','picking')])
            rec.countdone = self.env['kadouane.kadouane'].search_count([('state','=','done')])
            rec.counttransport = self.env['kadouane.kadouane'].search_count([('state','=','transport')])
            rec.countdouane = self.env['kadouane.kadouane'].search_count([('state','=','douane')])
            rec.countexp = self.env['kadouane.kadouane'].search_count([('stateShip','=','exp')])
            rec.countCloture = self.env['kadouane.kadouane'].search_count([('state','=','Cloture')])
            rec.countmainlevee = self.env['kadouane.kadouane'].search_count([('state','=','mainlevee')])
            rec.countFacturation = self.env['kadouane.kadouane'].search_count([('state','=','facturation')])
            rec.countArchive = self.env['kadouane.kadouane'].search_count([('state','=','archive'),('archiveclosed','=',False)])
            rec.countArchiveOK = self.env['kadouane.kadouane'].search_count([('state','=','archive'),('archiveclosed','=',True)])

    def _get_action(self, action_xmlid):
        action = self.env["ir.actions.actions"]._for_xml_id(action_xmlid)
        if self:
            action['display_name'] = self.display_name

        return action
